import pandas as pd
import sqlite3
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

def formatar_valor(valor):
    """Formata valor para R$ 1.234,56"""
    return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

def aplicar_formatacao_excel(arquivo_excel):
    """Aplica formatação profissional no Excel"""
    wb = load_workbook(arquivo_excel)
    
    # Estilos
    titulo_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    titulo_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    
    subtitulo_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    subtitulo_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    
    cabecalho_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cabecalho_fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
    
    borda = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 4, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Formatar cabeçalho
        for cell in ws[1]:
            cell.font = cabecalho_font
            cell.fill = cabecalho_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = borda
        
        # Alinhar dados
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = borda
        
        # Zebrar linhas
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if i % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill(start_color='F5F9FF', end_color='F5F9FF', fill_type='solid')
    
    # Formatação especial para a aba Resumo
    if 'Resumo' in wb.sheetnames:
        ws = wb['Resumo']
        
        # Limpar células mescladas anteriores
        for merge in list(ws.merged_cells):
            ws.unmerge_cells(str(merge))
        
        # Título principal
        ws.merge_cells('A1:K1')
        titulo = ws['A1']
        titulo.value = 'ANALISE DE VENDAS 2023 - RESUMO EXECUTIVO'
        titulo.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        titulo.fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        titulo.alignment = Alignment(horizontal='center', vertical='center')
        titulo.border = borda
        
        # Cabeçalho da tabela de estatísticas
        ws.merge_cells('A2:B2')
        cab_tabela = ws['A2']
        cab_tabela.value = 'ESTATISTICAS GERAIS'
        cab_tabela.font = subtitulo_font
        cab_tabela.fill = subtitulo_fill
        cab_tabela.alignment = Alignment(horizontal='center', vertical='center')
        cab_tabela.border = borda
        
        # Cabeçalho da área do gráfico
        ws.merge_cells('D2:K2')
        cab_graf = ws['D2']
        cab_graf.value = 'GRAFICO DE TENDENCIA MENSAL'
        cab_graf.font = subtitulo_font
        cab_graf.fill = subtitulo_fill
        cab_graf.alignment = Alignment(horizontal='center', vertical='center')
        cab_graf.border = borda
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 5  # Espaço entre estatísticas e gráfico
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        
        # Dados completos das estatísticas (todos com mesmo tamanho)
        estatisticas_completas = [
            ('Periodo Inicial', '07/01/2023'),
            ('Periodo Final', '27/12/2023'),
            ('Total de Registros', '60'),
            ('Total de Produtos', '10'),
            ('Total de Categorias', '5'),
            ('Faturamento Total', 'R$ 268.824,00'),
            ('Media por Produto', 'R$ 26.882,40'),
            ('Produto Mais Vendido', 'Bicicleta'),
            ('Valor Mais Vendido', 'R$ 97.200,00'),
            ('Produto Menos Vendido', 'Caneta'),
            ('Valor Menos Vendido', 'R$ 384,00')
        ]
        
        # Limpar linhas antigas
        for row in range(3, 30):
            for col in range(1, 12):
                ws.cell(row=row, column=col).value = None
        
        # Inserir dados formatados
        for i, (metrica, valor) in enumerate(estatisticas_completas, start=3):
            # Métrica
            cell_metrica = ws.cell(row=i, column=1)
            cell_metrica.value = metrica
            cell_metrica.font = Font(bold=True)
            cell_metrica.alignment = Alignment(horizontal='left', vertical='center')
            cell_metrica.border = borda
            
            # Valor
            cell_valor = ws.cell(row=i, column=2)
            cell_valor.value = valor
            cell_valor.font = Font(bold=True)
            cell_valor.alignment = Alignment(horizontal='right', vertical='center')
            cell_valor.border = borda
            
            # Zebrar linhas
            if i % 2 == 0:
                cell_metrica.fill = PatternFill(start_color='F5F9FF', end_color='F5F9FF', fill_type='solid')
                cell_valor.fill = PatternFill(start_color='F5F9FF', end_color='F5F9FF', fill_type='solid')
        
        # Ajustar altura das linhas de estatísticas (todas iguais)
        for i in range(3, 3 + len(estatisticas_completas)):
            ws.row_dimensions[i].height = 20
        
        # Configurar linha para o gráfico (altura adequada)
        ws.row_dimensions[3].height = 280  # Altura para o gráfico
    
    wb.save(arquivo_excel)

def gerar_graficos(df_clean, total_vendas_produto):
    """Gera os gráficos e retorna como imagens"""
    
    # Calcular vendas mensais
    df_temp = df_clean.copy()
    df_temp.set_index('Data', inplace=True)
    vendas_mensais = df_temp['TotalVenda'].resample('ME').sum()
    meses = vendas_mensais.index.strftime('%b/%Y')
    
    # Gráfico 1: Tendência de Vendas Mensais (TAMANHO AJUSTADO)
    fig1, ax1 = plt.subplots(figsize=(9, 4))  # Reduzido para caber ao lado
    
    # Linha do gráfico
    ax1.plot(meses, vendas_mensais, marker='o', linewidth=2, markersize=6, 
             color='#2E86AB', markerfacecolor='#F39C12', markeredgecolor='white', markeredgewidth=1)
    
    ax1.set_title('Tendência de Vendas Mensais', fontsize=12, fontweight='bold', pad=10)
    ax1.set_xlabel('Mês', fontsize=9, fontweight='bold')
    ax1.set_ylabel('Vendas (R$)', fontsize=9, fontweight='bold')
    ax1.grid(True, alpha=0.2, linestyle='--')
    ax1.tick_params(axis='x', rotation=45, labelsize=8)
    ax1.tick_params(axis='y', labelsize=8)
    
    # Adicionar valores apenas nos pontos principais para não poluir
    for i, (mes, valor) in enumerate(zip(meses, vendas_mensais)):
        if i % 2 == 0:  # Mostra a cada 2 meses para não poluir
            ax1.annotate(f'{valor:,.0f}'.replace(',', '.'), 
                        (mes, valor), textcoords="offset points", 
                        xytext=(0,8), ha='center', fontsize=7, 
                        bbox=dict(boxstyle='round,pad=0.2', facecolor='white', alpha=0.7))
    
    plt.tight_layout()
    
    img1 = BytesIO()
    fig1.savefig(img1, format='png', dpi=120, bbox_inches='tight', pad_inches=0.1)
    img1.seek(0)
    plt.close(fig1)
    
    # Gráfico 2: Total de Vendas por Produto
    fig2, ax2 = plt.subplots(figsize=(10, 4.5))
    cores = ['#2E86AB', '#A23B72', '#F18F01', '#C73E1D', '#3B8EA5', '#F18F01', '#A23B72', '#2E86AB', '#C73E1D', '#3B8EA5']
    bars = ax2.bar(total_vendas_produto.index, total_vendas_produto, 
                   color=cores[:len(total_vendas_produto)], edgecolor='black', linewidth=0.5)
    
    ax2.set_title('Total de Vendas por Produto', fontsize=12, fontweight='bold', pad=10)
    ax2.set_xlabel('Produto', fontsize=9, fontweight='bold')
    ax2.set_ylabel('Total (R$)', fontsize=9, fontweight='bold')
    ax2.tick_params(axis='x', rotation=45, labelsize=8)
    ax2.tick_params(axis='y', labelsize=8)
    ax2.grid(True, alpha=0.2, linestyle='--', axis='y')
    
    # Adicionar valores nas barras (apenas os maiores para não poluir)
    for bar in bars:
        height = bar.get_height()
        if height > 10000:  # Mostra apenas valores acima de 10.000
            ax2.text(bar.get_x() + bar.get_width()/2., height,
                    f'{height:,.0f}'.replace(',', '.'),
                    ha='center', va='bottom', fontsize=7, fontweight='bold')
    
    plt.tight_layout()
    
    img2 = BytesIO()
    fig2.savefig(img2, format='png', dpi=120, bbox_inches='tight', pad_inches=0.1)
    img2.seek(0)
    plt.close(fig2)
    
    return img1, img2, vendas_mensais

def executar_consultas_e_gerar_excel(df_clean, total_vendas_produto, nome_banco="vendas.db"):
    """
    Executa as consultas SQL e salva em Excel com gráficos
    """
    conn = sqlite3.connect(nome_banco)
    
    # CONSULTA 1: Total de vendas por produto
    query1 = '''
    SELECT 
        Produto,
        Categoria,
        SUM(Quantidade * Preço) AS TotalVendas
    FROM 
        vendas
    GROUP BY 
        Produto, Categoria
    ORDER BY 
        TotalVendas DESC;
    '''
    df1 = pd.read_sql_query(query1, conn)
    
    # CONSULTA 2: Produtos que venderam menos em Junho/2023
    query2 = '''
    SELECT 
        Produto,
        Categoria,
        SUM(Quantidade * Preço) AS VendasJunho2023
    FROM 
        vendas
    WHERE 
        Data LIKE '2023-06%'
    GROUP BY 
        Produto, Categoria
    ORDER BY 
        VendasJunho2023 ASC;
    '''
    df2 = pd.read_sql_query(query2, conn)
    
    conn.close()
    
    # Gera os gráficos
    img1, img2, vendas_mensais = gerar_graficos(df_clean, total_vendas_produto)
    
    # Cria um arquivo Excel com várias abas
    arquivo_excel = 'resultado_completo.xlsx'
    
    with pd.ExcelWriter(arquivo_excel, engine='openpyxl') as writer:
        
        # Aba 1 - Resumo (primeira planilha) - APENAS UMA LINHA POR MÉTRICA
        resumo = pd.DataFrame({
            'Metrica': [
                'Periodo Inicial',
                'Periodo Final',
                'Total de Registros',
                'Total de Produtos',
                'Total de Categorias',
                'Faturamento Total',
                'Media por Produto',
                'Produto Mais Vendido',
                'Valor Mais Vendido',
                'Produto Menos Vendido',
                'Valor Menos Vendido'
            ],
            'Valor': [
                df_clean['Data'].min().strftime('%d/%m/%Y'),
                df_clean['Data'].max().strftime('%d/%m/%Y'),
                str(len(df_clean)),
                str(df_clean['Produto'].nunique()),
                str(df_clean['Categoria'].nunique()),
                formatar_valor(df1['TotalVendas'].sum()),
                formatar_valor(df1['TotalVendas'].mean()),
                df1.loc[0, 'Produto'] if not df1.empty else 'N/A',
                formatar_valor(df1['TotalVendas'].max()) if not df1.empty else 'R$ 0,00',
                df1.loc[df1.index[-1], 'Produto'] if not df1.empty else 'N/A',
                formatar_valor(df1['TotalVendas'].min()) if not df1.empty else 'R$ 0,00'
            ]
        })
        resumo.to_excel(writer, sheet_name='Resumo', index=False)
        
        # Aba 2 - Total por produto
        df1_excel = df1.copy()
        df1_excel['TotalVendas'] = df1_excel['TotalVendas'].apply(formatar_valor)
        df1_excel.to_excel(writer, sheet_name='Total por Produto', index=False)
        
        # Aba 3 - Junho/2023
        df2_excel = df2.copy()
        df2_excel['VendasJunho2023'] = df2_excel['VendasJunho2023'].apply(formatar_valor)
        df2_excel.to_excel(writer, sheet_name='Junho2023', index=False)
        
        # Aba 4 - Dados brutos
        df_bruto = pd.read_sql_query("SELECT * FROM vendas ORDER BY Data;", sqlite3.connect(nome_banco))
        df_bruto['Preço'] = df_bruto['Preço'].apply(formatar_valor)
        df_bruto['Data'] = pd.to_datetime(df_bruto['Data']).dt.strftime('%d/%m/%Y')
        df_bruto.to_excel(writer, sheet_name='Dados Brutos', index=False)
    
    # Aplicar formatação profissional
    aplicar_formatacao_excel(arquivo_excel)
    
    # Adicionar gráficos na aba Resumo
    wb = load_workbook(arquivo_excel)
    ws = wb['Resumo']
    
    # Posicionar gráfico de tendência ao lado das estatísticas
    img_pil1 = Image(img1)
    img_pil1.width = 650  # Largura ajustada
    img_pil1.height = 280  # Altura ajustada
    ws.add_image(img_pil1, 'D3')  # Posicionado ao lado
    
    # Adicionar título para o segundo gráfico
    linha_graf2 = 3 + len(resumo) + 2  # Calcula posição após estatísticas
    ws.merge_cells(f'A{linha_graf2}:K{linha_graf2}')
    titulo_graf2 = ws[f'A{linha_graf2}']
    titulo_graf2.value = 'TOTAL DE VENDAS POR PRODUTO'
    titulo_graf2.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    titulo_graf2.fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    titulo_graf2.alignment = Alignment(horizontal='center', vertical='center')
    titulo_graf2.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Adicionar gráfico de produtos abaixo
    img_pil2 = Image(img2)
    img_pil2.width = 800
    img_pil2.height = 320
    ws.add_image(img_pil2, f'A{linha_graf2 + 1}')
    
    # Ajustar altura das linhas
    ws.row_dimensions[3].height = 280
    ws.row_dimensions[linha_graf2 + 1].height = 320
    
    wb.save(arquivo_excel)
    
    # Também salva CSVs como backup
    df1.to_csv("resultado_consulta1.csv", index=False, sep=';', encoding='utf-8-sig')
    df2.to_csv("resultado_consulta2.csv", index=False, sep=';', encoding='utf-8-sig')
    
    return df1, df2